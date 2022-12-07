import subprocess
import os
import re
import PyPDF2
import openpyxl as px

class ExcelPdfConverter:
  def __init__(self, original_excel_file: str, out_dir:str="/home/myuser/app/files/") -> None:
    if not os.path.isfile(original_excel_file):
      raise FileNotFoundError
    if not re.search(r'.xlsx$', original_excel_file):
      raise Exception("Please input .xlsx format file")
    self.__original_excel_file: str = original_excel_file
    self.__base_name: str = os.path.splitext(os.path.basename(original_excel_file))[0]
    self.__out_dir: str = out_dir

  def editExcel(self) -> None:
    wb = px.load_workbook(self.__original_excel_file)
    ws = wb.active
    ws["A2"].value = "001"
    ws["B2"].value = "山田太郎"
    ws["C2"].value = "18"
    wb.save(f"{self.__out_dir+self.__base_name}.xlsx")

  def convertPdf(self) -> None:
    cmd: list = []
    cmd.append("libreoffice")
    cmd.append("--headless")
    cmd.append("--nologo")
    cmd.append("--convert-to")
    cmd.append("pdf")
    cmd.append("--outdir")
    cmd.append(self.__out_dir)
    cmd.append(f"{self.__out_dir+self.__base_name}.xlsx")
    subprocess.run(" ".join(cmd), shell=True)

  def pdfCompress(self) -> None:
    writer = PyPDF2.PdfFileWriter()
    reader = PyPDF2.PdfFileReader(f"{self.__out_dir+self.__base_name}.pdf")
    for i in range(reader.getNumPages()):
      page = reader.getPage(i)
      page.compressContentStreams()
      writer.addPage(page)
    with open(f"{self.__out_dir+self.__base_name}-compressed.pdf", "wb") as fp:
      writer.write(fp)

if __name__ == "__main__":
  converter = ExcelPdfConverter("sample.xlsx")
  converter.editExcel()
  converter.convertPdf()
  converter.pdfCompress()